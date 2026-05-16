"""Reports Service for generating Excel and PDF exports"""
import io
from datetime import datetime, date
from typing import Optional, List, Literal
from sqlalchemy.orm import Session
from sqlalchemy import and_, func

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from app.models.inventory import InventoryBatch
from app.models.order import Order, OrderStatus, PaymentStatus
from app.models.order_item import OrderItem
from app.models.prescription import Prescription
from app.models.medicine import Medicine
from app.models.store import Store


class ReportsService:
    """Service for generating export reports in Excel and PDF formats"""
    
    def __init__(self, db: Session):
        self.db = db
    
    # ============= Excel Helpers =============
    
    def _create_excel_workbook(self, title: str, headers: List[str], data: List[List]) -> io.BytesIO:
        """Create a styled Excel workbook"""
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet name limit
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = max(
                len(str(ws.cell(row=1, column=col).value or "")),
                max((len(str(ws.cell(row=r, column=col).value or "")) for r in range(2, len(data) + 2)), default=0)
            )
            ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 50)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    # ============= PDF Helpers =============
    
    def _create_pdf_document(self, title: str, headers: List[str], data: List[List], 
                             subtitle: str = None) -> io.BytesIO:
        """Create a styled PDF document"""
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=12,
            alignment=1  # Center
        )
        elements.append(Paragraph(title, title_style))
        
        # Subtitle (date range, etc.)
        if subtitle:
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Normal'],
                fontSize=10,
                spaceAfter=20,
                alignment=1,
                textColor=colors.grey
            )
            elements.append(Paragraph(subtitle, subtitle_style))
        
        elements.append(Spacer(1, 12))
        
        # Table
        table_data = [headers] + data
        
        # Calculate column widths based on content
        col_count = len(headers)
        available_width = 7.5 * inch
        col_width = available_width / col_count
        
        table = Table(table_data, colWidths=[col_width] * col_count)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
        ]))
        elements.append(table)
        
        # Footer with generation timestamp
        elements.append(Spacer(1, 20))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=2  # Right
        )
        elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", footer_style))
        
        doc.build(elements)
        output.seek(0)
        return output
    
    # ============= Inventory Report =============
    
    def export_inventory_report(
        self,
        store_id: Optional[int] = None,
        format: Literal["excel", "pdf"] = "excel"
    ) -> io.BytesIO:
        """Export inventory report"""
        query = self.db.query(InventoryBatch).filter(InventoryBatch.inactive == False)
        
        if store_id:
            query = query.filter(InventoryBatch.store_id == store_id)
        
        batches = query.all()
        
        headers = ["Batch Number", "Medicine", "Store", "Quantity", "Unit Price", 
                   "Expiry Date", "Manufacturer", "Status"]
        
        data = []
        for batch in batches:
            medicine_name = batch.medicine.brand_name if batch.medicine else "N/A"
            store_name = batch.store.name if batch.store else "N/A"
            status = "Low Stock" if batch.quantity <= batch.reorder_level else "OK"
            if batch.expiry_date and batch.expiry_date < date.today():
                status = "Expired"
            
            data.append([
                batch.batch_number,
                medicine_name,
                store_name,
                batch.quantity,
                f"₹{batch.unit_price:.2f}",
                batch.expiry_date.strftime("%Y-%m-%d") if batch.expiry_date else "N/A",
                batch.manufacturer or "N/A",
                status
            ])
        
        title = "Inventory Report"
        subtitle = f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        
        if format == "pdf":
            return self._create_pdf_document(title, headers, data, subtitle)
        return self._create_excel_workbook(title, headers, data)
    
    # ============= Orders Report =============
    
    def export_orders_report(
        self,
        store_id: Optional[int] = None,
        date_from: Optional[date] = None,
        date_to: Optional[date] = None,
        status: Optional[str] = None,
        format: Literal["excel", "pdf"] = "excel"
    ) -> io.BytesIO:
        """Export orders report"""
        query = self.db.query(Order)
        
        if store_id:
            query = query.filter(Order.store_id == store_id)
        if date_from:
            query = query.filter(Order.created_at >= date_from)
        if date_to:
            query = query.filter(Order.created_at <= date_to)
        if status:
            query = query.filter(Order.status == status)
        
        orders = query.order_by(Order.created_at.desc()).all()
        
        headers = ["Order #", "Date", "Customer", "Store", "Items", 
                   "Total (₹)", "Status", "Payment"]
        
        data = []
        for order in orders:
            customer_name = order.customer.full_name if order.customer else "Walk-in"
            store_name = order.store.name if order.store else "N/A"
            
            data.append([
                order.order_number,
                order.created_at.strftime("%Y-%m-%d %H:%M"),
                customer_name,
                store_name,
                len(order.items),
                f"{order.total_amount:.2f}",
                order.status.value,
                order.payment_status.value
            ])
        
        title = "Orders Report"
        subtitle = f"Period: {date_from or 'All'} to {date_to or 'Present'}"
        
        if format == "pdf":
            return self._create_pdf_document(title, headers, data, subtitle)
        return self._create_excel_workbook(title, headers, data)
    
    # ============= Prescriptions Report =============
    
    def export_prescriptions_report(
        self,
        store_id: Optional[int] = None,
        date_from: Optional[date] = None,
        date_to: Optional[date] = None,
        format: Literal["excel", "pdf"] = "pdf"
    ) -> io.BytesIO:
        """Export prescriptions report (PDF only for compliance)"""
        query = self.db.query(Prescription)
        
        if store_id:
            query = query.filter(Prescription.store_id == store_id)
        if date_from:
            query = query.filter(Prescription.created_at >= date_from)
        if date_to:
            query = query.filter(Prescription.created_at <= date_to)
        
        prescriptions = query.order_by(Prescription.created_at.desc()).all()
        
        headers = ["Rx ID", "Date", "Patient", "Doctor", "Store", "Status", "Verified By"]
        
        data = []
        for rx in prescriptions:
            patient_name = rx.customer.full_name if rx.customer else "N/A"
            store_name = rx.store.name if rx.store else "N/A"
            verified_by = rx.verified_by_user.full_name if rx.verified_by_user else "Pending"
            
            data.append([
                f"RX-{rx.id}",
                rx.created_at.strftime("%Y-%m-%d"),
                patient_name,
                rx.doctor_name or "N/A",
                store_name,
                rx.status.value,
                verified_by
            ])
        
        title = "Prescriptions Report"
        subtitle = f"Period: {date_from or 'All'} to {date_to or 'Present'}"
        
        # Always PDF for prescriptions (regulatory compliance)
        return self._create_pdf_document(title, headers, data, subtitle)
    
    # ============= Sales Report =============
    
    def export_sales_report(
        self,
        store_id: Optional[int] = None,
        date_from: Optional[date] = None,
        date_to: Optional[date] = None,
        format: Literal["excel", "pdf"] = "excel"
    ) -> io.BytesIO:
        """Export sales report with summary"""
        query = self.db.query(Order).filter(Order.payment_status == PaymentStatus.PAID)
        
        if store_id:
            query = query.filter(Order.store_id == store_id)
        if date_from:
            query = query.filter(Order.created_at >= date_from)
        if date_to:
            query = query.filter(Order.created_at <= date_to)
        
        orders = query.order_by(Order.created_at.desc()).all()
        
        headers = ["Order #", "Date", "Store", "Subtotal (₹)", "Tax (₹)", 
                   "Discount (₹)", "Total (₹)", "Payment Method"]
        
        data = []
        total_sales = 0
        total_tax = 0
        total_discount = 0
        
        for order in orders:
            store_name = order.store.name if order.store else "N/A"
            payment_method = order.payment_method.value if order.payment_method else "N/A"
            
            data.append([
                order.order_number,
                order.created_at.strftime("%Y-%m-%d"),
                store_name,
                f"{order.subtotal:.2f}",
                f"{order.tax_amount:.2f}",
                f"{order.discount_amount:.2f}",
                f"{order.total_amount:.2f}",
                payment_method
            ])
            
            total_sales += order.total_amount
            total_tax += order.tax_amount
            total_discount += order.discount_amount
        
        # Add summary row
        data.append(["", "", "TOTAL", "", f"{total_tax:.2f}", 
                     f"{total_discount:.2f}", f"{total_sales:.2f}", ""])
        
        title = "Sales Report"
        subtitle = f"Period: {date_from or 'All'} to {date_to or 'Present'} | Total Sales: ₹{total_sales:.2f}"
        
        if format == "pdf":
            return self._create_pdf_document(title, headers, data, subtitle)
        return self._create_excel_workbook(title, headers, data)
