"""
Generate DETAILED Test Results Excel Report with comprehensive test information
"""
import pandas as pd
from app.db.session import SessionLocal
from app.models.test_history import TestRun, TestResult
from datetime import datetime
import re

# Connect to database
db = SessionLocal()

# Get all test results with their run information
results = db.query(TestResult).join(TestRun).all()

# Prepare detailed data for Excel
detailed_test_data = []

for result in results:
    # Extract test class and method
    parts = result.nodeid.split("::")
    test_file = parts[0] if len(parts) > 0 else "Unknown"
    test_class = parts[1] if len(parts) > 1 else "N/A"
    test_method = parts[2] if len(parts) > 2 else parts[-1]
    
    # Parse error details
    error_type = "N/A"
    error_line = "N/A"
    error_detail = "N/A"
    
    if result.error_message:
        # Try to extract error type
        if "assert" in result.error_message.lower():
            error_type = "Assertion Error"
        elif "attributeerror" in result.error_message.lower():
            error_type = "Attribute Error"
        elif "typeerror" in result.error_message.lower():
            error_type = "Type Error"
        elif "keyerror" in result.error_message.lower():
            error_type = "Key Error"
        elif "valueerror" in result.error_message.lower():
            error_type = "Value Error"
        else:
            error_type = "Runtime Error"
        
        # Try to extract line number
        line_match = re.search(r'line (\d+)', result.error_message)
        if line_match:
            error_line = line_match.group(1)
        
        # First 300 chars of error for detail
        error_detail = result.error_message[:300]
    
    # Determine test category based on test name
    category = "Unit Test"
    if "integration" in result.nodeid.lower() or "flow" in test_method.lower():
        category = "Integration Test"
    elif "test_create" in test_method.lower():
        category = "CRUD - Create"
    elif "test_read" in test_method.lower() or "test_get" in test_method.lower():
        category = "CRUD - Read"
    elif "test_update" in test_method.lower():
        category = "CRUD - Update"
    elif "test_delete" in test_method.lower():
        category = "CRUD - Delete"
    elif "lifecycle" in test_method.lower():
        category = "Full Lifecycle"
    elif "auth" in test_method.lower() or "login" in test_method.lower():
        category = "Authentication"
    
    # Status details
    status_emoji = "✅" if result.status == "passed" else "❌" if result.status == "failed" else "⏭️"
    
    detailed_test_data.append({
        "Project Code": result.test_run.project_code,
        "Module Code": result.module_code or "Unknown",
        "Module Name": result.module_code.replace("_", " ").title() if result.module_code else "Unknown",
        "Category": category,
        "Test File": test_file.replace("tests/", "").replace("tests\\", ""),
        "Test Class": test_class,
        "Test Method": test_method,
        "Status": f"{status_emoji} {result.status.upper()}",
        "Duration (s)": round(result.duration, 4),
        "Performance": "Fast" if result.duration < 0.5 else "Normal" if result.duration < 2.0 else "Slow",
        "Error Type": error_type,
        "Error Line": error_line,
        "Error Summary": error_detail,
        "Full Error": result.error_message or "No errors",
        "Test Run ID": result.test_run.id,
        "Run Timestamp": result.test_run.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
        "Environment": result.test_run.environment,
        "Full Test Path": result.nodeid,
    })

db.close()

# Create DataFrame
df = pd.DataFrame(detailed_test_data)

# Sort by Module Code, Category, Status
df = df.sort_values(['Module Code', 'Category', 'Status'])

# Create comprehensive summary
module_summary = []
for module in sorted(df['Module Code'].unique()):
    module_df = df[df['Module Code'] == module]
    total = len(module_df)
    passed = len(module_df[module_df['Status'].str.contains('PASSED')])
    failed = len(module_df[module_df['Status'].str.contains('FAILED')])
    skipped = len(module_df[module_df['Status'].str.contains('SKIPPED')])
    avg_duration = module_df['Duration (s)'].mean()
    max_duration = module_df['Duration (s)'].max()
    min_duration = module_df['Duration (s)'].min()
    project_code = module_df['Project Code'].iloc[0] if len(module_df) > 0 else "N/A"
    
    module_summary.append({
        "Project Code": project_code,
        "Module Code": module,
        "Module Name": module.replace("_", " ").title() if module else "Unknown",
        "Total Tests": total,
        "✅ Passed": passed,
        "❌ Failed": failed,
        "⏭️ Skipped": skipped,
        "Pass Rate %": round((passed / total * 100) if total > 0 else 0, 1),
        "Avg Duration (s)": round(avg_duration, 3),
        "Min Duration (s)": round(min_duration, 3),
        "Max Duration (s)": round(max_duration, 3),
        "Total Time (s)": round(module_df['Duration (s)'].sum(), 3),
    })

summary_df = pd.DataFrame(module_summary)

# Category summary
category_summary = []
for category in sorted(df['Category'].unique()):
    cat_df = df[df['Category'] == category]
    total = len(cat_df)
    passed = len(cat_df[cat_df['Status'].str.contains('PASSED')])
    
    category_summary.append({
        "Test Category": category,
        "Total Tests": total,
        "✅ Passed": passed,
        "❌ Failed": total - passed,
        "Pass Rate %": round((passed / total * 100) if total > 0 else 0, 1),
        "Avg Duration (s)": round(cat_df['Duration (s)'].mean(), 3),
    })

category_df = pd.DataFrame(category_summary)

# Failed tests details
failed_df = df[df['Status'].str.contains('FAILED')].copy()
failed_df = failed_df[['Project Code', 'Module Code', 'Module Name', 'Test Method', 'Error Type', 'Error Line', 'Error Summary', 'Duration (s)', 'Run Timestamp']]

# Performance analysis
perf_df = df[['Project Code', 'Module Code', 'Module Name', 'Test Method', 'Duration (s)', 'Performance', 'Status']].copy()
perf_df = perf_df.sort_values('Duration (s)', ascending=False)

# Create Excel file with multiple sheets
output_path = "d:/projects/meat-backed/Detailed_Test_Results.xlsx"

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    # Executive Summary
    summary_df.to_excel(writer, sheet_name='📊 Module Summary', index=False)
    
    # Category Breakdown
    category_df.to_excel(writer, sheet_name='📁 Category Summary', index=False)
    
    # Failed Tests Only
    if len(failed_df) > 0:
        failed_df.to_excel(writer, sheet_name='❌ Failed Tests', index=False)
    
    # Performance Analysis
    perf_df.to_excel(writer, sheet_name='⚡ Performance', index=False)
    
    # All Detailed Results
    df.to_excel(writer, sheet_name='📋 All Test Details', index=False)
    
    # Individual sheets per module with full details
    for module in sorted(df['Module Code'].unique()):
        module_df = df[df['Module Code'] == module]
        sheet_name = f"🔍 {module}"[:31]  # Excel sheet name limit
        module_df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Format all sheets
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    for sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]
        
        # Format header
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Set row height for header
        worksheet.row_dimensions[1].height = 30
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max(max_length + 3, 12), 120)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Apply conditional formatting to status and performance
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            for cell in row:
                # Wrap text for error columns
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                
                # Color code status
                if cell.value and isinstance(cell.value, str):
                    if "✅ PASSED" in cell.value:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(color="006100", bold=True)
                    elif "❌ FAILED" in cell.value:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006", bold=True)
                    elif "⏭️ SKIPPED" in cell.value:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        cell.font = Font(color="9C5700", bold=True)
                    
                    # Performance indicators
                    if cell.value == "Fast":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(color="006100")
                    elif cell.value == "Slow":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006")
        
        # Freeze header row
        worksheet.freeze_panes = "A2"
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
            for cell in row:
                cell.border = thin_border

# Print detailed summary
print(f"\n{'='*70}")
print(f"✅ DETAILED Test Results Excel Created Successfully!")
print(f"{'='*70}")
print(f"\n📁 File Location: {output_path}")
print(f"\n📊 Overall Statistics:")
print(f"  • Total Test Cases: {len(df)}")
print(f"  • Project: {df['Project Code'].iloc[0] if len(df) > 0 else 'N/A'}")
print(f"  • Modules Tested: {len(df['Module Code'].unique())}")
print(f"  • Test Categories: {len(df['Category'].unique())}")
print(f"  • ✅ Passed: {len(df[df['Status'].str.contains('PASSED')])}")
print(f"  • ❌ Failed: {len(df[df['Status'].str.contains('FAILED')])}")
print(f"  • ⏭️ Skipped: {len(df[df['Status'].str.contains('SKIPPED')])}")
print(f"  • Overall Pass Rate: {round((len(df[df['Status'].str.contains('PASSED')]) / len(df) * 100), 1)}%")
print(f"  • Total Execution Time: {round(df['Duration (s)'].sum(), 2)}s")
print(f"  • Average Test Duration: {round(df['Duration (s)'].mean(), 3)}s")

print(f"\n📋 Excel Sheets Created:")
print(f"  1. 📊 Module Summary - Overview by module")
print(f"  2. 📁 Category Summary - Breakdown by test type")
if len(failed_df) > 0:
    print(f"  3. ❌ Failed Tests - Detailed failure analysis")
print(f"  4. ⚡ Performance - Duration analysis")
print(f"  5. 📋 All Test Details - Complete test information")

print(f"\n🔍 Module-Specific Sheets:")
for idx, module in enumerate(sorted(df['Module Code'].unique()), start=6):
    count = len(df[df['Module Code'] == module])
    passed = len(df[(df['Module Code'] == module) & (df['Status'].str.contains('PASSED'))])
    print(f"  {idx}. 🔍 {module} - {count} tests ({passed} passed)")

print(f"\n{'='*70}")
print(f"📝 Each test includes:")
print(f"  ✓ Project Code & Module Code for tracking")
print(f"  ✓ Module Name & Category classification")
print(f"  ✓ Test file, class, and method names")
print(f"  ✓ Execution status with visual indicators")
print(f"  ✓ Duration and performance rating")
print(f"  ✓ Error type, line number, and summary")
print(f"  ✓ Full error trace for debugging")
print(f"  ✓ Test run metadata and timestamp")
print(f"{'='*70}\n")
