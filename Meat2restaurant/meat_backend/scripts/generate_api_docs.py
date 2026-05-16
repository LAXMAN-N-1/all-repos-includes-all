"""
Generate comprehensive API documentation Excel file
"""
import pandas as pd
from pathlib import Path
import re

# Define all endpoints with details
endpoints_data = [
    # Authentication Endpoints
    {
        "Module": "Authentication",
        "Method": "POST",
        "Endpoint": "/api/v1/auth/register",
        "Description": "Register a new user account with email and password. Creates user record in database with hashed password.",
        "Flow": "Client sends email, password, full_name → System validates uniqueness → Hashes password → Creates user → Returns user details",
        "Auth Required": "No",
        "Role": "Public"
    },
    {
        "Module": "Authentication",
        "Method": "POST",
        "Endpoint": "/api/v1/auth/login",
        "Description": "Authenticate user credentials and generate JWT access token for subsequent API calls.",
        "Flow": "Client sends username/password → System validates credentials → Generates JWT token → Returns access_token and token_type",
        "Auth Required": "No",
        "Role": "Public"
    },
    {
        "Module": "Authentication",
        "Method": "POST",
        "Endpoint": "/api/v1/auth/refresh",
        "Description": "Refresh an expired JWT token to maintain user session without re-authentication.",
        "Flow": "Client sends expired token → System validates token → Generates new JWT → Returns new access_token",
        "Auth Required": "Yes",
        "Role": "Any authenticated user"
    },
    
    # User Management
    {
        "Module": "Users",
        "Method": "GET",
        "Endpoint": "/api/v1/users/me",
        "Description": "Retrieve current authenticated user's profile information including email, role, and permissions.",
        "Flow": "Client sends request with JWT → System extracts user from token → Returns user profile data",
        "Auth Required": "Yes",
        "Role": "Any authenticated user"
    },
    {
        "Module": "Users",
        "Method": "GET",
        "Endpoint": "/api/v1/users/",
        "Description": "List all users in the system with pagination support. Admin-only endpoint for user management.",
        "Flow": "Admin sends request → System queries users table → Returns paginated list of users",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    {
        "Module": "Users",
        "Method": "POST",
        "Endpoint": "/api/v1/users/",
        "Description": "Create a new user account with specified role and permissions. Admin can set custom roles.",
        "Flow": "Admin sends user data → System validates → Creates user with hashed password → Returns created user",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    {
        "Module": "Users",
        "Method": "PUT",
        "Endpoint": "/api/v1/users/{user_id}",
        "Description": "Update existing user's information including email, role, or active status.",
        "Flow": "Admin sends user_id and update data → System validates → Updates user record → Returns updated user",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    {
        "Module": "Users",
        "Method": "DELETE",
        "Endpoint": "/api/v1/users/{user_id}",
        "Description": "Soft delete or permanently remove a user account from the system.",
        "Flow": "Admin sends user_id → System validates → Marks user as inactive or deletes → Returns confirmation",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    
    # Products
    {
        "Module": "Products",
        "Method": "GET",
        "Endpoint": "/api/v1/products/",
        "Description": "Retrieve paginated list of all products with filtering by category, price range, and availability.",
        "Flow": "Client sends query params → System filters products → Returns list with pagination metadata",
        "Auth Required": "No",
        "Role": "Public"
    },
    {
        "Module": "Products",
        "Method": "GET",
        "Endpoint": "/api/v1/products/{product_id}",
        "Description": "Get detailed information about a specific product including pricing, stock, and specifications.",
        "Flow": "Client sends product_id → System queries product → Returns full product details with related data",
        "Auth Required": "No",
        "Role": "Public"
    },
    {
        "Module": "Products",
        "Method": "POST",
        "Endpoint": "/api/v1/products/",
        "Description": "Create new product listing with details like name, price, SKU, category, and stock quantity.",
        "Flow": "Admin sends product data → System validates → Creates product record → Returns created product with ID",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    {
        "Module": "Products",
        "Method": "PUT",
        "Endpoint": "/api/v1/products/{product_id}",
        "Description": "Update existing product information including price, stock, description, or category.",
        "Flow": "Admin sends product_id and updates → System validates → Updates product → Returns updated product",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    {
        "Module": "Products",
        "Method": "DELETE",
        "Endpoint": "/api/v1/products/{product_id}",
        "Description": "Remove product from catalog or mark as inactive to prevent new orders.",
        "Flow": "Admin sends product_id → System checks dependencies → Soft deletes product → Returns confirmation",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    
    # Catalog
    {
        "Module": "Catalog",
        "Method": "GET",
        "Endpoint": "/api/v1/catalog/categories",
        "Description": "List all product categories with hierarchical structure for navigation and filtering.",
        "Flow": "Client requests categories → System retrieves category tree → Returns nested category structure",
        "Auth Required": "No",
        "Role": "Public"
    },
    {
        "Module": "Catalog",
        "Method": "POST",
        "Endpoint": "/api/v1/catalog/categories",
        "Description": "Create new product category with optional parent category for hierarchical organization.",
        "Flow": "Admin sends category data → System validates parent_id → Creates category → Returns created category",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    {
        "Module": "Catalog",
        "Method": "PUT",
        "Endpoint": "/api/v1/catalog/categories/{category_id}",
        "Description": "Update category details or reorganize hierarchy by changing parent category.",
        "Flow": "Admin sends updates → System validates circular references → Updates category → Returns updated data",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    {
        "Module": "Catalog",
        "Method": "DELETE",
        "Endpoint": "/api/v1/catalog/categories/{category_id}",
        "Description": "Remove category from system after checking no products are assigned to it.",
        "Flow": "Admin sends category_id → System checks product dependencies → Deletes if safe → Returns confirmation",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    {
        "Module": "Catalog",
        "Method": "GET",
        "Endpoint": "/api/v1/catalog/attributes",
        "Description": "Retrieve product attributes (e.g., size, color, cut type) for filtering and specifications.",
        "Flow": "Client requests attributes → System queries active attributes → Returns list of attributes",
        "Auth Required": "No",
        "Role": "Public"
    },
    {
        "Module": "Catalog",
        "Method": "POST",
        "Endpoint": "/api/v1/catalog/attributes",
        "Description": "Create new product attribute definition for use in product specifications.",
        "Flow": "Admin sends attribute data → System validates uniqueness → Creates attribute → Returns created attribute",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    {
        "Module": "Catalog",
        "Method": "PUT",
        "Endpoint": "/api/v1/catalog/attributes/{attribute_id}",
        "Description": "Update attribute name or active status for product specification management.",
        "Flow": "Admin sends attribute_id and updates → System validates → Updates attribute → Returns updated data",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    {
        "Module": "Catalog",
        "Method": "DELETE",
        "Endpoint": "/api/v1/catalog/attributes/{attribute_id}",
        "Description": "Remove attribute definition from system for catalog cleanup.",
        "Flow": "Admin sends attribute_id → System validates → Deletes attribute → Returns confirmation",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    
    # Customers
    {
        "Module": "Customers",
        "Method": "GET",
        "Endpoint": "/api/v1/customers/",
        "Description": "List all customer accounts with filtering by status, type, and registration date.",
        "Flow": "Staff requests customers → System queries with filters → Returns paginated customer list",
        "Auth Required": "Yes",
        "Role": "Staff/Admin"
    },
    {
        "Module": "Customers",
        "Method": "GET",
        "Endpoint": "/api/v1/customers/{customer_id}",
        "Description": "Get detailed customer profile including contact info, order history, and membership status.",
        "Flow": "Staff sends customer_id → System retrieves customer data → Returns full customer profile",
        "Auth Required": "Yes",
        "Role": "Staff/Admin or Self"
    },
    {
        "Module": "Customers",
        "Method": "POST",
        "Endpoint": "/api/v1/customers/",
        "Description": "Register new customer account with business details for B2B platform access.",
        "Flow": "User sends registration data → System validates → Creates customer record → Returns customer with ID",
        "Auth Required": "No",
        "Role": "Public"
    },
    {
        "Module": "Customers",
        "Method": "PUT",
        "Endpoint": "/api/v1/customers/{customer_id}",
        "Description": "Update customer information including business name, contact details, or delivery addresses.",
        "Flow": "Customer/Admin sends updates → System validates → Updates customer → Returns updated profile",
        "Auth Required": "Yes",
        "Role": "Staff/Admin or Self"
    },
    {
        "Module": "Customers",
        "Method": "GET",
        "Endpoint": "/api/v1/customers/membership-plans",
        "Description": "List available membership tiers with pricing, benefits, and duration details.",
        "Flow": "Client requests plans → System retrieves active plans → Returns list of membership options",
        "Auth Required": "No",
        "Role": "Public"
    },
    {
        "Module": "Customers",
        "Method": "POST",
        "Endpoint": "/api/v1/customers/membership-plans",
        "Description": "Create new membership tier with custom pricing and benefits for customer retention.",
        "Flow": "Admin sends plan data → System validates → Creates plan → Returns created membership plan",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    {
        "Module": "Customers",
        "Method": "POST",
        "Endpoint": "/api/v1/customers/{customer_id}/membership",
        "Description": "Assign membership plan to customer with automatic expiry date calculation.",
        "Flow": "Admin/Customer assigns plan → System calculates end_date → Creates membership → Returns membership details",
        "Auth Required": "Yes",
        "Role": "Admin or Self"
    },
    {
        "Module": "Customers",
        "Method": "GET",
        "Endpoint": "/api/v1/customers/{customer_id}/membership",
        "Description": "Retrieve customer's current active membership status and expiry information.",
        "Flow": "Client sends customer_id → System queries membership → Returns active membership or 404",
        "Auth Required": "Yes",
        "Role": "Staff/Admin or Self"
    },
    {
        "Module": "Customers",
        "Method": "DELETE",
        "Endpoint": "/api/v1/customers/{customer_id}/membership",
        "Description": "Cancel customer's membership subscription and revoke associated benefits.",
        "Flow": "Admin/Customer cancels → System marks membership inactive → Returns cancellation confirmation",
        "Auth Required": "Yes",
        "Role": "Admin or Self"
    },
    
    # Orders
    {
        "Module": "Orders",
        "Method": "GET",
        "Endpoint": "/api/v1/orders/",
        "Description": "List orders with filtering by status, customer, date range, and payment status.",
        "Flow": "User requests orders → System filters by permissions → Returns paginated order list",
        "Auth Required": "Yes",
        "Role": "Staff/Admin or Customer (own orders)"
    },
    {
        "Module": "Orders",
        "Method": "GET",
        "Endpoint": "/api/v1/orders/{order_id}",
        "Description": "Get complete order details including line items, pricing, shipping, and payment status.",
        "Flow": "User sends order_id → System validates access → Returns full order with related data",
        "Auth Required": "Yes",
        "Role": "Staff/Admin or Order Owner"
    },
    {
        "Module": "Orders",
        "Method": "POST",
        "Endpoint": "/api/v1/orders/",
        "Description": "Create new order with products, quantities, shipping address, and payment method selection.",
        "Flow": "Customer sends order data → System validates stock → Calculates totals → Creates order → Returns order with ID",
        "Auth Required": "Yes",
        "Role": "Customer/Partner"
    },
    {
        "Module": "Orders",
        "Method": "PUT",
        "Endpoint": "/api/v1/orders/{order_id}",
        "Description": "Update order status (pending, processing, shipped, delivered) or modify order details before fulfillment.",
        "Flow": "Staff sends order_id and updates → System validates state transition → Updates order → Returns updated order",
        "Auth Required": "Yes",
        "Role": "Staff/Admin"
    },
    {
        "Module": "Orders",
        "Method": "DELETE",
        "Endpoint": "/api/v1/orders/{order_id}",
        "Description": "Cancel order if not yet shipped and process refund if payment was made.",
        "Flow": "Customer/Admin cancels → System checks status → Reverses stock → Initiates refund → Returns confirmation",
        "Auth Required": "Yes",
        "Role": "Staff/Admin or Order Owner"
    },
    
    # Invoices
    {
        "Module": "Invoices",
        "Method": "GET",
        "Endpoint": "/api/v1/invoices/",
        "Description": "List invoices with filtering by customer, status, due date, and payment status.",
        "Flow": "User requests invoices → System filters by permissions → Returns paginated invoice list",
        "Auth Required": "Yes",
        "Role": "Staff/Admin or Customer (own invoices)"
    },
    {
        "Module": "Invoices",
        "Method": "GET",
        "Endpoint": "/api/v1/invoices/{invoice_id}",
        "Description": "Retrieve detailed invoice with line items, tax calculations, and payment history.",
        "Flow": "User sends invoice_id → System validates access → Returns invoice with PDF URL",
        "Auth Required": "Yes",
        "Role": "Staff/Admin or Invoice Owner"
    },
    {
        "Module": "Invoices",
        "Method": "POST",
        "Endpoint": "/api/v1/invoices/",
        "Description": "Generate invoice for order with automatic PDF creation and email notification to customer.",
        "Flow": "System/Admin creates invoice → Generates PDF → Sends email → Returns invoice with PDF URL",
        "Auth Required": "Yes",
        "Role": "Staff/Admin"
    },
    {
        "Module": "Invoices",
        "Method": "PUT",
        "Endpoint": "/api/v1/invoices/{invoice_id}",
        "Description": "Update invoice status (draft, sent, paid, overdue) or modify payment terms.",
        "Flow": "Admin sends updates → System validates → Updates invoice → Triggers notifications → Returns updated invoice",
        "Auth Required": "Yes",
        "Role": "Staff/Admin"
    },
    
    # Payments
    {
        "Module": "Payments",
        "Method": "POST",
        "Endpoint": "/api/v1/payments/",
        "Description": "Process payment for order or invoice using configured payment gateway (Stripe, PayPal, etc.).",
        "Flow": "Customer submits payment → System validates → Processes via gateway → Updates order status → Returns payment confirmation",
        "Auth Required": "Yes",
        "Role": "Customer/Partner"
    },
    {
        "Module": "Payments",
        "Method": "GET",
        "Endpoint": "/api/v1/payments/{payment_id}",
        "Description": "Retrieve payment transaction details including status, amount, and gateway response.",
        "Flow": "User sends payment_id → System validates access → Returns payment details with transaction info",
        "Auth Required": "Yes",
        "Role": "Staff/Admin or Payment Owner"
    },
    {
        "Module": "Payments",
        "Method": "POST",
        "Endpoint": "/api/v1/payments/webhook",
        "Description": "Handle payment gateway webhooks for async payment status updates (success, failure, refund).",
        "Flow": "Gateway sends webhook → System validates signature → Updates payment status → Triggers order updates",
        "Auth Required": "No (Signature verified)",
        "Role": "Payment Gateway"
    },
    
    # Promotions
    {
        "Module": "Promotions",
        "Method": "GET",
        "Endpoint": "/api/v1/promotions/",
        "Description": "List active promotions and discount codes with validity dates and usage limits.",
        "Flow": "Client requests promotions → System filters active promotions → Returns available discounts",
        "Auth Required": "No",
        "Role": "Public"
    },
    {
        "Module": "Promotions",
        "Method": "POST",
        "Endpoint": "/api/v1/promotions/",
        "Description": "Create promotional campaign with discount percentage, code, validity period, and usage limits.",
        "Flow": "Admin sends promotion data → System validates → Creates promotion → Returns created promotion",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    {
        "Module": "Promotions",
        "Method": "POST",
        "Endpoint": "/api/v1/promotions/validate",
        "Description": "Validate promotion code during checkout to apply discount and check usage limits.",
        "Flow": "Customer sends code → System checks validity, usage, expiry → Returns discount amount or error",
        "Auth Required": "Yes",
        "Role": "Customer/Partner"
    },
    {
        "Module": "Promotions",
        "Method": "PUT",
        "Endpoint": "/api/v1/promotions/{promotion_id}",
        "Description": "Update promotion details including discount amount, validity dates, or active status.",
        "Flow": "Admin sends updates → System validates → Updates promotion → Returns updated promotion",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    
    # Locations
    {
        "Module": "Locations",
        "Method": "GET",
        "Endpoint": "/api/v1/locations/",
        "Description": "List delivery locations with service availability and delivery zone information.",
        "Flow": "Client requests locations → System retrieves active locations → Returns location list",
        "Auth Required": "No",
        "Role": "Public"
    },
    {
        "Module": "Locations",
        "Method": "POST",
        "Endpoint": "/api/v1/locations/",
        "Description": "Add new delivery location with address, coordinates, and service area definition.",
        "Flow": "Admin sends location data → System validates → Creates location → Returns created location",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    
    # Settings
    {
        "Module": "Settings",
        "Method": "GET",
        "Endpoint": "/api/v1/settings/shipping",
        "Description": "Retrieve available shipping methods with pricing, delivery times, and service areas.",
        "Flow": "Client requests shipping options → System retrieves methods → Returns shipping methods",
        "Auth Required": "No",
        "Role": "Public"
    },
    {
        "Module": "Settings",
        "Method": "POST",
        "Endpoint": "/api/v1/settings/shipping",
        "Description": "Create new shipping method with pricing rules and delivery time estimates.",
        "Flow": "Admin sends method data → System validates → Creates shipping method → Returns created method",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    {
        "Module": "Settings",
        "Method": "GET",
        "Endpoint": "/api/v1/settings/config",
        "Description": "Get system configuration including business hours, contact info, and feature flags.",
        "Flow": "Client requests config → System retrieves settings → Returns configuration object",
        "Auth Required": "No",
        "Role": "Public"
    },
    
    # Sales
    {
        "Module": "Sales",
        "Method": "GET",
        "Endpoint": "/api/v1/sales/gift-cards",
        "Description": "List gift cards with balance, expiry date, and usage history for customer loyalty.",
        "Flow": "Staff requests gift cards → System queries cards → Returns list with balances",
        "Auth Required": "Yes",
        "Role": "Staff/Admin"
    },
    {
        "Module": "Sales",
        "Method": "POST",
        "Endpoint": "/api/v1/sales/gift-cards",
        "Description": "Create new gift card with initial balance and expiry date for customer purchase.",
        "Flow": "Admin sends card data → System generates unique code → Creates card → Returns card details",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    {
        "Module": "Sales",
        "Method": "POST",
        "Endpoint": "/api/v1/sales/gift-cards/redeem",
        "Description": "Redeem gift card during checkout by deducting amount from balance and applying to order.",
        "Flow": "Customer sends code and amount → System validates balance → Deducts amount → Returns new balance",
        "Auth Required": "Yes",
        "Role": "Customer/Partner"
    },
    
    # Analytics
    {
        "Module": "Analytics",
        "Method": "GET",
        "Endpoint": "/api/v1/analytics/sales",
        "Description": "Retrieve sales analytics with revenue trends, top products, and customer segments.",
        "Flow": "Admin requests analytics → System aggregates sales data → Returns charts and metrics",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    {
        "Module": "Analytics",
        "Method": "GET",
        "Endpoint": "/api/v1/analytics/revenue-by-category",
        "Description": "Get revenue breakdown by product category for business intelligence and inventory planning.",
        "Flow": "Admin requests category revenue → System groups by category → Returns revenue per category",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    {
        "Module": "Analytics",
        "Method": "GET",
        "Endpoint": "/api/v1/analytics/event-analytics",
        "Description": "Track user events and behavior patterns for conversion optimization and UX improvements.",
        "Flow": "Admin requests events → System aggregates event data → Returns event metrics and trends",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    
    # Reports
    {
        "Module": "Reports",
        "Method": "GET",
        "Endpoint": "/api/v1/reports/inventory",
        "Description": "Generate inventory report with stock levels, low stock alerts, and reorder suggestions.",
        "Flow": "Admin requests report → System calculates stock metrics → Returns inventory report",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    {
        "Module": "Reports",
        "Method": "GET",
        "Endpoint": "/api/v1/reports/sales-summary",
        "Description": "Generate sales summary report for specified date range with revenue and order statistics.",
        "Flow": "Admin sends date range → System aggregates sales → Returns summary with totals",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    
    # WhatsApp Integration
    {
        "Module": "WhatsApp",
        "Method": "POST",
        "Endpoint": "/api/v1/whatsapp/webhook",
        "Description": "Handle incoming WhatsApp messages for order placement and customer support automation.",
        "Flow": "WhatsApp sends message → System parses intent → Processes command → Sends response via WhatsApp",
        "Auth Required": "No (Webhook verified)",
        "Role": "WhatsApp Service"
    },
    {
        "Module": "WhatsApp",
        "Method": "POST",
        "Endpoint": "/api/v1/whatsapp/send",
        "Description": "Send WhatsApp notifications to customers for order updates, promotions, and alerts.",
        "Flow": "System triggers notification → Formats message → Sends via WhatsApp API → Returns delivery status",
        "Auth Required": "Yes",
        "Role": "System/Admin"
    },
    
    # CMS
    {
        "Module": "CMS",
        "Method": "GET",
        "Endpoint": "/api/v1/cms/pages",
        "Description": "Retrieve CMS pages for about, terms, privacy policy, and other static content.",
        "Flow": "Client requests pages → System queries CMS → Returns page list with content",
        "Auth Required": "No",
        "Role": "Public"
    },
    {
        "Module": "CMS",
        "Method": "POST",
        "Endpoint": "/api/v1/cms/pages",
        "Description": "Create new CMS page with rich text content for website content management.",
        "Flow": "Admin sends page data → System validates → Creates page → Returns created page",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    {
        "Module": "CMS",
        "Method": "PUT",
        "Endpoint": "/api/v1/cms/pages/{page_id}",
        "Description": "Update CMS page content, title, or publish status for website updates.",
        "Flow": "Admin sends updates → System validates → Updates page → Returns updated page",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    
    # Partner Pricing
    {
        "Module": "Partner Pricing",
        "Method": "GET",
        "Endpoint": "/api/v1/partner-pricing/",
        "Description": "Retrieve custom pricing tiers for B2B partners based on volume and relationship.",
        "Flow": "Partner requests pricing → System retrieves partner tier → Returns custom pricing",
        "Auth Required": "Yes",
        "Role": "Partner/Admin"
    },
    {
        "Module": "Partner Pricing",
        "Method": "POST",
        "Endpoint": "/api/v1/partner-pricing/",
        "Description": "Create custom pricing agreement for specific partner with volume discounts.",
        "Flow": "Admin sends pricing data → System validates → Creates pricing tier → Returns pricing agreement",
        "Auth Required": "Yes",
        "Role": "Admin/Superuser"
    },
    
    # Test History
    {
        "Module": "Test History",
        "Method": "GET",
        "Endpoint": "/api/v1/test-history/runs",
        "Description": "Retrieve test execution history with pass/fail statistics and module coverage.",
        "Flow": "Developer requests history → System queries test_runs → Returns test execution records",
        "Auth Required": "Yes",
        "Role": "Developer/Admin"
    },
    {
        "Module": "Test History",
        "Method": "GET",
        "Endpoint": "/api/v1/test-history/runs/{run_id}",
        "Description": "Get detailed test run results including individual test outcomes and error messages.",
        "Flow": "Developer sends run_id → System retrieves run with results → Returns detailed test data",
        "Auth Required": "Yes",
        "Role": "Developer/Admin"
    },
]

# Create DataFrame
df = pd.DataFrame(endpoints_data)

# Reorder columns for better readability
df = df[["Module", "Method", "Endpoint", "Auth Required", "Role", "Description", "Flow"]]

# Create Excel file with formatting
output_path = "d:/projects/meat-backed/API_Documentation.xlsx"

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='API Endpoints', index=False)
    
    # Get the worksheet
    worksheet = writer.sheets['API Endpoints']
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 18  # Module
    worksheet.column_dimensions['B'].width = 8   # Method
    worksheet.column_dimensions['C'].width = 45  # Endpoint
    worksheet.column_dimensions['D'].width = 12  # Auth Required
    worksheet.column_dimensions['E'].width = 20  # Role
    worksheet.column_dimensions['F'].width = 80  # Description
    worksheet.column_dimensions['G'].width = 100 # Flow
    
    # Format header row
    from openpyxl.styles import Font, PatternFill, Alignment
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Format data rows
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            
        # Color code by method
        method_cell = row[1]
        if method_cell.value == "GET":
            method_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        elif method_cell.value == "POST":
            method_cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        elif method_cell.value == "PUT":
            method_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        elif method_cell.value == "DELETE":
            method_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    
    # Freeze header row
    worksheet.freeze_panes = "A2"

print(f"✅ API Documentation Excel created successfully!")
print(f"📁 Location: {output_path}")
print(f"📊 Total Endpoints Documented: {len(endpoints_data)}")
